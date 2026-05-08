"""Run the NMR-agent image-to-signal surrogate for one RTM mask pair.

The preferred input is one RTM interface image like the NMR-agent training
samples under Data/inputs/interface_images. The script applies the same
red/yellow phase detection and white-border crop used during dataset building.

For debugging, two CSV masks can still be supplied instead:
  - solid mask, matching NMR-agent channel 0
  - pore/liquid mask, matching NMR-agent channel 1

The script writes a simple two-column Excel workbook:
  time_s, pred_signal

That workbook is intentionally compatible with the existing RTSPHEM
T2_process inversion bridge.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import numpy as np
import torch
from openpyxl import Workbook
from PIL import Image


def _json_result(payload: dict[str, Any]) -> None:
    print("RESULT_JSON=" + json.dumps(payload, ensure_ascii=False, separators=(",", ":")))


def _infer_nmr_agent_root(model_path: Path) -> Path:
    # <root>/runs/<run_id>/latest_model.pt
    try:
        return model_path.resolve().parents[2]
    except IndexError as exc:
        raise ValueError(f"Cannot infer NMR-agent root from model path: {model_path}") from exc


def _resolve_dataset_path(checkpoint: dict[str, Any], dataset_arg: str | None, nmr_agent_root: Path) -> Path:
    raw = dataset_arg or checkpoint.get("dataset_path")
    if not raw:
        raise ValueError("Dataset path was not provided and is absent from the checkpoint.")
    path = Path(str(raw))
    if not path.is_absolute():
        path = nmr_agent_root / path
    return path


def _load_mask_csv(path: Path, resolution: int) -> np.ndarray:
    mask = np.loadtxt(path, delimiter=",", dtype=np.float32)
    if mask.ndim != 2:
        raise ValueError(f"Mask must be 2-D: {path}")
    mask = np.clip(mask, 0.0, 1.0)
    image = Image.fromarray((mask * 255.0).astype(np.uint8))
    resized = image.resize((resolution, resolution), Image.Resampling.BILINEAR)
    return (np.asarray(resized, dtype=np.float32) / 255.0).astype(np.float32)


def _load_masks_from_interface_image(path: Path, resolution: int) -> np.ndarray:
    from nmr_surrogate.image_signal_data import colored_bbox, phase_masks_from_image  # noqa: WPS433

    rgb = np.asarray(Image.open(path).convert("RGB"))
    bbox = colored_bbox(rgb, padding=0)
    return phase_masks_from_image(path, bbox, resolution)


def _condition_vector(length_x_axis: float, length_y_axis: float, resolution: int, dataset: dict[str, np.ndarray]) -> np.ndarray:
    raw = np.asarray(
        [
            np.log(length_x_axis),
            np.log(length_y_axis),
            np.log(length_x_axis / resolution),
            np.log(length_y_axis / resolution),
            length_x_axis / length_y_axis,
        ],
        dtype=np.float32,
    )
    mean = np.asarray(dataset["condition_mean"], dtype=np.float32)
    std = np.asarray(dataset["condition_std"], dtype=np.float32)
    std = np.where(std < 1e-6, 1.0, std).astype(np.float32)
    return ((raw - mean) / std).astype(np.float32)


def _write_decay_workbook(path: Path, time_s: np.ndarray, signal: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "surrogate_decay"
    sheet.cell(row=1, column=1, value="time_s")
    sheet.cell(row=1, column=2, value="pred_signal")
    for row_idx, (time_value, signal_value) in enumerate(zip(time_s, signal), start=2):
        sheet.cell(row=row_idx, column=1, value=float(time_value))
        sheet.cell(row=row_idx, column=2, value=float(signal_value))
    workbook.save(path)


def _write_decay_csv(path: Path, time_s: np.ndarray, signal: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        handle.write("time_s,pred_signal\n")
        for time_value, signal_value in zip(time_s, signal):
            handle.write(f"{float(time_value):.9g},{float(signal_value):.9g}\n")


def run(args: argparse.Namespace) -> dict[str, Any]:
    model_path = Path(args.model_path)
    if not model_path.exists():
        raise FileNotFoundError(f"Surrogate model not found: {model_path}")

    nmr_agent_root = Path(args.nmr_agent_root) if args.nmr_agent_root else _infer_nmr_agent_root(model_path)
    if str(nmr_agent_root) not in sys.path:
        sys.path.insert(0, str(nmr_agent_root))

    from nmr_surrogate.unet_signal import UNetSignalRegressor  # noqa: WPS433

    if args.device == "auto":
        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    else:
        device = torch.device(args.device)

    checkpoint = torch.load(model_path, map_location=device)
    dataset_path = _resolve_dataset_path(checkpoint, args.dataset_path, nmr_agent_root)
    if not dataset_path.exists():
        raise FileNotFoundError(f"Surrogate dataset metadata not found: {dataset_path}")
    loaded = np.load(dataset_path, allow_pickle=False)
    dataset = {key: loaded[key] for key in loaded.files}

    time_s = np.asarray(dataset["signal_time_s"], dtype=np.float32)
    resolution = int(args.resolution or dataset["masks"].shape[-1])
    train_config = checkpoint.get("train_config", {})
    condition_dim = int(checkpoint.get("condition_dim", len(dataset.get("condition_mean", []))))
    model = UNetSignalRegressor(
        signal_length=len(time_s),
        condition_dim=condition_dim,
        base_channels=int(train_config.get("base_channels", 24)),
        dropout=float(train_config.get("dropout", 0.1)),
    ).to(device)
    model.load_state_dict(checkpoint["model_state_dict"])
    model.eval()

    if args.interface_image:
        masks = _load_masks_from_interface_image(Path(args.interface_image), resolution)[None, :, :, :]
    else:
        if not args.solid_mask_csv or not args.pore_mask_csv:
            raise ValueError("Provide --interface-image, or both --solid-mask-csv and --pore-mask-csv.")
        solid = _load_mask_csv(Path(args.solid_mask_csv), resolution)
        pore = _load_mask_csv(Path(args.pore_mask_csv), resolution)
        masks = np.stack([solid, pore], axis=0)[None, :, :, :]
    conditions = _condition_vector(args.length_x_axis, args.length_y_axis, resolution, dataset)[None, :]

    with torch.no_grad():
        pred = model(
            torch.from_numpy(masks).float().to(device),
            torch.from_numpy(conditions).float().to(device),
        )["signal"][0]
    signal = pred.detach().cpu().numpy().astype(np.float32)

    output_excel = Path(args.output_excel)
    _write_decay_workbook(output_excel, time_s, signal)
    output_csv = Path(args.output_csv) if args.output_csv else output_excel.with_suffix(".csv")
    _write_decay_csv(output_csv, time_s, signal)

    metadata_path = output_excel.with_suffix(".json")
    metadata = {
        "model_path": str(model_path),
        "dataset_path": str(dataset_path),
        "nmr_agent_root": str(nmr_agent_root),
        "device": str(device),
        "resolution": resolution,
        "length_x_axis": float(args.length_x_axis),
        "length_y_axis": float(args.length_y_axis),
        "interface_image": str(args.interface_image) if args.interface_image else "",
        "signal_points": int(len(signal)),
        "output_excel": str(output_excel),
        "output_csv": str(output_csv),
    }
    metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "success": True,
        "output_excel": str(output_excel),
        "output_csv": str(output_csv),
        "metadata_json": str(metadata_path),
        "signal_points": int(len(signal)),
        "signal_min": float(np.nanmin(signal)),
        "signal_max": float(np.nanmax(signal)),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Predict an NMR relaxation curve with the image surrogate.")
    parser.add_argument("--interface-image", default="")
    parser.add_argument("--solid-mask-csv", default="")
    parser.add_argument("--pore-mask-csv", default="")
    parser.add_argument("--length-x-axis", type=float, required=True)
    parser.add_argument("--length-y-axis", type=float, required=True)
    parser.add_argument("--model-path", required=True)
    parser.add_argument("--dataset-path", default="")
    parser.add_argument("--nmr-agent-root", default="")
    parser.add_argument("--output-excel", required=True)
    parser.add_argument("--output-csv", default="")
    parser.add_argument("--resolution", type=int, default=0)
    parser.add_argument("--device", default="auto")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        _json_result(run(args))
        return 0
    except Exception as exc:  # noqa: BLE001
        _json_result({"success": False, "error": str(exc)})
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
